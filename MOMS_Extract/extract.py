"""
MOMS Checklist Extractor — Universal v4
=========================================
Production-grade extraction of MOMS checklist data for any WI number.

Key improvements in v4:
  - Empty rows suppressed (unfilled template slots not written)
  - Date filter: only records from --from_year onwards (default 2021)
  - File splitting: separate Excel sheet per year, plus combined sheet
  - File size guard: splits into multiple files if > MAX_ROWS_PER_FILE
  - Dynamic header resolution for all table types
  - Graceful handling if output file is already open
  - Progress indicator during formatting
  - FieldLabel truncation for long radio step descriptions
  - Retry logic on file save

Usage:
    python extract.py --wi SIG/WR/PMW/0007
    python extract.py --wi SIG/WR/PMW/0007 --limit 500
    python extract.py --wi SIG/WR/PMW/0007 --limit 0
    python extract.py --wi SIG/WR/PMW/0007 --from_year 2021
    python extract.py --wi SIG/WR/PMW/0007 --output C:/Reports

Author: SMRT Technical Operations
"""

import argparse
import logging
import os
import re
import sys
import time
from datetime import datetime
from math import ceil

import pandas as pd
import pyodbc
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ── CONFIG ────────────────────────────────────────────────────────────────────
DB_CONFIG = {
    "driver":   "SQL Server",
    "server":   "SQLRAILAGLP1",
    "port":     "9762",
    "database": "SMRT_MOMS_DB",
}

# Maximum rows per Excel file before splitting into multiple files
MAX_ROWS_PER_FILE = 200_000

# Maximum FieldLabel length (radio step descriptions can be very long)
MAX_LABEL_LEN = 200

# ── CONTROL TYPE CLASSIFICATION ───────────────────────────────────────────────
CTRL_CATEGORY = {
    "Radio":    "result",
    "DropDown": "selection",
    "Textbox":  "measurement",
    "TextArea": "narrative",
    "Remark":   "remark",
    "CheckBox": "confirmation",
    "Date":     "datetime",
    "Header":   "layout",
    "Label":    "layout",
    "SignPad":  "signature",
    "Initials": "signature",
}

INPUT_CATEGORIES = {
    "result", "selection", "measurement",
    "narrative", "remark", "confirmation", "datetime"
}

# Controls where an empty value is still meaningful to keep
# (e.g. a radio with no result = unanswered step — analytically significant)
KEEP_IF_EMPTY = {"result"}

# Excel colours
COLORS = {
    "header_bg":  "1F4E79",
    "section_bg": "2E75B6",
    "year_bg":    "E9EFF7",
    "ok_bg":      "E2EFDA",
    "nok_bg":     "FCE4D6",
    "na_bg":      "FFF2CC",
    "alt_row":    "F2F2F2",
    "white":      "FFFFFF",
    "warn":       "C00000",
    "grey_text":  "595959",
}


# ── DATABASE ──────────────────────────────────────────────────────────────────
def get_connection():
    conn_str = (
        f"DRIVER={{{DB_CONFIG['driver']}}};"
        f"SERVER={DB_CONFIG['server']},{DB_CONFIG['port']};"
        f"DATABASE={DB_CONFIG['database']};"
        f"Trusted_Connection=yes;"
    )
    try:
        conn = pyodbc.connect(conn_str, timeout=30)
        log.info(f"Connected  →  {DB_CONFIG['database']} @ {DB_CONFIG['server']}")
        return conn
    except pyodbc.Error as e:
        log.error(f"DB connection failed: {e}")
        sys.exit(1)


def fetch_records(wi_number: str, limit: int, from_year: int) -> list:
    conn   = get_connection()
    cursor = conn.cursor()
    top    = f"TOP {limit}" if limit > 0 else ""
    query  = f"""
        SELECT {top}
            ID, Checklist, WorkOrderID,
            ChecklistTemplateID, ChecklistTemplateLookup,
            Description, Created, CreatedBy,
            Modified, ModifiedBy, FilledBy, VersionNo, IsDeleted
        FROM   Checklist
        WHERE  IsDeleted = 0
          AND  ChecklistTemplateID = ?
          AND  YEAR(Created) >= ?
        ORDER  BY ID DESC
    """
    cursor.execute(query, (wi_number, from_year))
    rows = cursor.fetchall()
    conn.close()
    log.info(
        f"Fetched {len(rows):,} records  "
        f"(WI: {wi_number}, from: {from_year}, limit: {limit or 'ALL'})"
    )
    return rows


# ── XML SECTION MODEL ─────────────────────────────────────────────────────────
def build_section_model(section: ET.Element) -> dict:
    """
    Index all controls in a section by row and column.
    Also detects header rows and builds col→header label mapping.
    """
    row_map        = {}   # row_num: [ctrl_dict, ...]
    col_map        = {}   # col_num: [ctrl_dict, ...]
    header_rows    = []   # list of row numbers that are pure-header rows
    col_header_map = {}   # (header_row, col): label

    for ctrl in section.findall(".//ControlBase"):
        r   = int(ctrl.findtext("Row",        "0"))
        c   = int(ctrl.findtext("Column",     "0"))
        cs  = int(ctrl.findtext("ColumnSpan", "1"))
        t   =     ctrl.findtext("Type",       "")
        lbl = (   ctrl.findtext("Label",      "") or "")\
                  .strip()\
                  .replace("#lbh#", " ")\
                  .replace("\n", " ")\
                  .replace("\t", " ")\
                  .strip()
        dat = (ctrl.findtext("Data", "") or "").strip()

        d = {
            "row":       r,
            "col":       c,
            "col_span":  cs,
            "type":      t,
            "label":     lbl,
            "data":      dat,
            "group":     (ctrl.findtext("GroupName",    "") or "").strip(),
            "name":      (ctrl.findtext("n",            "") or "").strip(),
            "mandatory": (ctrl.findtext("Mandatory",    "") or "").strip(),
            "dd_items":  (ctrl.findtext("DropDownItems","") or "").strip(),
            "min_val":   (ctrl.findtext("Min",          "") or "").strip(),
            "max_val":   (ctrl.findtext("Max",          "") or "").strip(),
        }

        row_map.setdefault(r, []).append(d)
        for offset in range(cs):
            col_map.setdefault(c + offset, []).append(d)

    # Detect header rows and build col→label index
    for r, controls in sorted(row_map.items()):
        non_empty = [c for c in controls if c["label"] or c["data"]]
        if not non_empty:
            continue
        if (all(c["type"] == "Header" for c in non_empty)
                and any(c["label"] for c in non_empty)
                and len(non_empty) >= 2):
            header_rows.append(r)
            for c in non_empty:
                for offset in range(c["col_span"]):
                    col_header_map[(r, c["col"] + offset)] = c["label"]

    return {
        "row_map":       row_map,
        "col_map":       col_map,
        "header_rows":   sorted(header_rows),
        "col_header_map":col_header_map,
    }


# ── DYNAMIC LABEL RESOLUTION ──────────────────────────────────────────────────
_SKIP_HEADERS = frozenset(["WI TITLE", "WR NO.", "WI NO.", "PAGE"])


def resolve_column_header(model: dict, ctrl_row: int, ctrl_col: int) -> str:
    """
    Scan upward from ctrl_row to find the nearest header row covering ctrl_col.
    Skips generic WI title headers.
    """
    above = [r for r in model["header_rows"] if r < ctrl_row]
    for hdr_row in sorted(above, reverse=True):
        label = model["col_header_map"].get((hdr_row, ctrl_col), "")
        if label and not any(s in label.upper() for s in _SKIP_HEADERS):
            return re.sub(r"\s+", " ", label).strip()
    return ""


def resolve_row_label(model: dict, ctrl_row: int, ctrl_col: int) -> str:
    """
    Scan left on the same row for the nearest Label or Header control.
    """
    left = [
        c for c in model["row_map"].get(ctrl_row, [])
        if c["type"] in ("Label", "Header")
        and c["col"] < ctrl_col
        and c["label"]
    ]
    if not left:
        return ""
    return re.sub(r"\s+", " ", max(left, key=lambda c: c["col"])["label"]).strip()


def build_field_name(col_header: str, row_label: str, ctrl_label: str,
                     max_len: int = MAX_LABEL_LEN) -> str:
    """
    Combine col_header and row_label into a meaningful field name.
    Truncates to max_len to prevent excessively long labels.
    """
    parts = []

    if col_header:
        clean = col_header[:80] if len(col_header) > 80 else col_header
        parts.append(clean)

    if row_label:
        clean = re.sub(r"\s+", " ", row_label).strip()
        # Only add if not already contained in col_header
        if clean and clean.lower() not in (col_header or "").lower():
            parts.append(clean)

    if not parts and ctrl_label:
        parts.append(ctrl_label)

    result = " > ".join(parts)
    if len(result) > max_len:
        result = result[:max_len - 3] + "..."
    return result


# ── RADIO RESULT ──────────────────────────────────────────────────────────────
def radio_result(row_controls: list) -> str:
    radio_map = {c["col"]: c["data"] for c in row_controls if c["type"] == "Radio"}
    if radio_map.get(6) == "true": return "OK"
    if radio_map.get(7) == "true": return "NOK"
    if radio_map.get(8) == "true": return "NA"
    return ""


# ── METADATA DETECTION ────────────────────────────────────────────────────────
def detect_wi_metadata(root: ET.Element) -> dict:
    """Extract form-level metadata. Flexible — works for any WI layout."""
    meta = {
        "WI_Title":    "",
        "WI_Number":   "",
        "WR_Number":   "",
        "MaintOrder":  "",
        "Station":     "",
        "EquipmentID": "",
        "Multimeter":  "",
    }

    flat = []
    for ctrl in root.findall(".//ControlBase"):
        flat.append({
            "row":   int(ctrl.findtext("Row",    "0")),
            "col":   int(ctrl.findtext("Column", "0")),
            "type":  ctrl.findtext("Type",  ""),
            "label": (ctrl.findtext("Label", "") or "").strip().replace("#lbh#", " "),
            "data":  (ctrl.findtext("Data",  "") or "").strip(),
        })

    def adjacent_input(fragment: str) -> str:
        matches = [
            c for c in flat
            if c["type"] == "Header"
            and fragment.lower() in c["label"].lower()
        ]
        if not matches:
            return ""
        hdr = matches[0]
        candidates = sorted(
            [c for c in flat
             if c["row"] == hdr["row"]
             and c["type"] in ("Textbox", "DropDown", "Date")
             and c["col"] > hdr["col"]],
            key=lambda x: x["col"]
        )
        return candidates[0]["data"] if candidates else ""

    # WI Title: longest Header label in first 3 rows, ignoring meta labels
    title_skip = {"WI TITLE", "WI NO", "WR NO", "PAGE"}
    candidates = [
        c for c in flat
        if c["type"] == "Header"
        and c["row"] <= 3
        and len(c["label"]) > 15
        and not any(s in c["label"].upper() for s in title_skip)
    ]
    if candidates:
        meta["WI_Title"] = max(candidates, key=lambda x: len(x["label"]))["label"]

    # WI/WR Number: Header with slash pattern
    for c in flat:
        if c["type"] == "Header" and c["row"] <= 5:
            lbl = c["label"]
            if re.search(r"[A-Z]{2,}/[A-Z]{2,}/", lbl):
                if "WI" in lbl and not meta["WI_Number"]:
                    meta["WI_Number"] = lbl
                if "WR" in lbl and not meta["WR_Number"]:
                    meta["WR_Number"] = lbl

    meta["MaintOrder"]  = adjacent_input("Maintenance Order")
    meta["Station"]     = adjacent_input("Station")
    meta["EquipmentID"] = adjacent_input("Equipment ID")
    meta["Multimeter"]  = adjacent_input("Multimeter")

    return meta


# ── SUBSECTION CONTEXT ────────────────────────────────────────────────────────
def update_context(ctx: dict, row_controls: list) -> bool:
    """
    Detect and update subsection context (17.x, 17.x.x patterns).
    Returns True if this row is a context-only row (no data to emit).
    """
    non_empty = [c for c in row_controls if c["label"] or c["data"]]
    if not non_empty:
        return False
    if not all(c["type"] in ("Header", "Label") for c in non_empty):
        return False

    col0 = [c["label"] for c in non_empty if c["col"] == 0 and c["label"]]
    col2 = [c["label"] for c in non_empty if c["col"] == 2 and c["label"]]

    if not col0:
        return False

    lbl0 = col0[0]
    col2_span = next((c["col_span"] for c in non_empty if c["col"] == 2), 1)

    if re.match(r"^\d+\.\d+$", lbl0) and col2:
        ctx.update({
            "subsection_no":   lbl0,
            "subsection_name": col2[0],
            "subitem_no":      "",
            "subitem_name":    "",
        })
        return True

    if re.match(r"^\d+\.\d+\.\d+$", lbl0) and col2_span >= 5:
        ctx.update({
            "subitem_no":   lbl0,
            "subitem_name": col2[0] if col2 else "",
        })
        return True

    return False


# ── CORE PARSER ───────────────────────────────────────────────────────────────
def parse_record(row) -> dict:
    """
    Parse one MOMS record into structured step and raw control rows.
    Only emits rows where value is filled OR category is in KEEP_IF_EMPTY.
    Returns: {'steps': [...], 'raw': [...], 'error': str|None}
    """
    out = {"steps": [], "raw": [], "error": None}
    xml_str = row.Checklist

    if not xml_str or not str(xml_str).strip():
        out["error"] = "Empty XML"
        return out

    try:
        root = ET.fromstring(xml_str)
    except ET.ParseError as e:
        out["error"] = f"Parse error: {e}"
        return out

    base = {
        "ID":          row.ID,
        "WorkOrderID": str(row.WorkOrderID or "").strip(),
        "TemplateID":  str(row.ChecklistTemplateID or "").strip(),
        "Template":    str(row.ChecklistTemplateLookup or "").strip(),
        "Created":     row.Created,
        "CreatedBy":   str(row.CreatedBy or "").strip(),
        "FilledBy":    str(row.FilledBy or "").strip(),
        "VersionNo":   row.VersionNo,
        **detect_wi_metadata(root),
    }

    for sec_idx, section in enumerate(root.findall(".//Section")):
        sec_name = (section.findtext("n", "") or "").strip() or f"Section_{sec_idx}"

        # Skip sections with no analytical controls at all
        type_counts = {}
        for c in section.findall(".//ControlBase"):
            t = c.findtext("Type", "")
            type_counts[t] = type_counts.get(t, 0) + 1
        if not any(CTRL_CATEGORY.get(t, "") in INPUT_CATEGORIES
                   for t in type_counts):
            continue

        model = build_section_model(section)

        ctx = {
            "subsection_no":   "",
            "subsection_name": "",
            "subitem_no":      "",
            "subitem_name":    "",
        }

        for row_num in sorted(model["row_map"].keys()):
            row_controls = model["row_map"][row_num]

            if update_context(ctx, row_controls):
                continue

            inputs = [
                c for c in row_controls
                if CTRL_CATEGORY.get(c["type"], "") in INPUT_CATEGORIES
            ]
            if not inputs:
                continue

            for ctrl in inputs:
                ctrl_type = ctrl["type"]
                category  = CTRL_CATEGORY.get(ctrl_type, "other")
                raw_value = ctrl["data"]

                # ── Normalise value ───────────────────────────────────────────
                if ctrl_type == "Radio":
                    if ctrl["col"] != 6:
                        continue  # Only emit once per radio group
                    value = radio_result(row_controls)

                elif ctrl_type == "CheckBox":
                    value = ("YES" if raw_value == "true"
                             else "NO" if raw_value == "false"
                             else "")

                else:
                    value = raw_value

                # ── Filter: skip empty rows unless analytically significant ───
                is_empty = not value or str(value).strip() == ""
                if is_empty and category not in KEEP_IF_EMPTY:
                    continue

                # ── Resolve labels ────────────────────────────────────────────
                col_header = resolve_column_header(model, row_num, ctrl["col"])
                row_label  = resolve_row_label(model, row_num, ctrl["col"])

                # For radio steps, the best field name is the step description
                if ctrl_type == "Radio":
                    step_desc = next(
                        (c["label"] for c in row_controls
                         if c["type"] == "Label" and c["col"] == 2),
                        ""
                    )
                    field_name = build_field_name(
                        col_header, step_desc or row_label, ctrl["label"]
                    )
                else:
                    field_name = build_field_name(col_header, row_label, ctrl["label"])

                step_no = next(
                    (c["label"] for c in row_controls
                     if c["type"] == "Label" and c["col"] == 0),
                    ""
                )

                inline_remark = (
                    next(
                        (c["data"] for c in row_controls
                         if c["type"] in ("Remark", "TextArea") and c["col"] == 9),
                        ""
                    ) if ctrl_type == "Radio" else ""
                )

                # ── Checklist_Steps row ───────────────────────────────────────
                out["steps"].append({
                    **base,
                    "SectionIdx":     sec_idx,
                    "SectionName":    sec_name,
                    "SubSectionNo":   ctx["subsection_no"],
                    "SubSectionName": ctx["subsection_name"],
                    "SubItemNo":      ctx["subitem_no"],
                    "SubItemName":    ctx["subitem_name"],
                    "StepNo":         step_no,
                    "FieldLabel":     field_name,
                    "ColHeader":      col_header,
                    "RowLabel":       row_label,
                    "CtrlType":       ctrl_type,
                    "Category":       category,
                    "Value":          value,
                    "IsFilled":       not is_empty,
                    "IsNOK":          value == "NOK",
                    "Remark":         inline_remark,
                    "HasRemark":      bool(inline_remark),
                    "DropDownOpts":   ctrl["dd_items"],
                    "Mandatory":      ctrl["mandatory"],
                })

                # ── Raw_Controls row (full fidelity) ──────────────────────────
                out["raw"].append({
                    **base,
                    "SectionIdx":   sec_idx,
                    "SectionName":  sec_name,
                    "GridRow":      row_num,
                    "GridCol":      ctrl["col"],
                    "ColSpan":      ctrl["col_span"],
                    "CtrlType":     ctrl_type,
                    "Category":     category,
                    "ColHeader":    col_header,
                    "RowLabel":     row_label,
                    "FieldLabel":   field_name,
                    "CtrlLabel":    ctrl["label"],
                    "CtrlName":     ctrl["name"],
                    "GroupName":    ctrl["group"],
                    "RawValue":     raw_value,
                    "Value":        value,
                    "DropDownOpts": ctrl["dd_items"],
                    "Mandatory":    ctrl["mandatory"],
                    "MinVal":       ctrl["min_val"],
                    "MaxVal":       ctrl["max_val"],
                })

    return out


# ── DATAFRAME POST-PROCESSING ─────────────────────────────────────────────────
def postprocess(df_steps: pd.DataFrame, df_raw: pd.DataFrame):
    """Clean and enrich DataFrames after parsing."""
    if df_steps.empty:
        return df_steps, df_raw

    # Parse Created to datetime
    df_steps["Created"] = pd.to_datetime(df_steps["Created"], errors="coerce")
    df_steps["Year"]    = df_steps["Created"].dt.year
    df_steps["Month"]   = df_steps["Created"].dt.to_period("M").astype(str)

    if not df_raw.empty:
        df_raw["Created"] = pd.to_datetime(df_raw["Created"], errors="coerce")
        df_raw["Year"]    = df_raw["Created"].dt.year

    # Remove duplicate step rows (same record + same grid position)
    before = len(df_steps)
    df_steps = df_steps.drop_duplicates(
        subset=["ID", "SectionIdx", "FieldLabel", "CtrlType", "Value"],
        keep="first"
    )
    dupes_removed = before - len(df_steps)
    if dupes_removed:
        log.info(f"Removed {dupes_removed:,} duplicate step rows")

    # Sort for readability
    df_steps = df_steps.sort_values(
        ["Created", "ID", "SectionIdx"],
        ascending=[False, False, True]
    ).reset_index(drop=True)

    return df_steps, df_raw


# ── SUMMARY ───────────────────────────────────────────────────────────────────
def build_summary(df_steps: pd.DataFrame, wi_number: str, limit: int,
                  from_year: int, record_count: int, error_count: int,
                  output_files: list) -> pd.DataFrame:

    if df_steps.empty:
        filled = nok = ok = na = total_steps = 0
        ctrl_dist = {}; cat_dist = {}; year_dist = {}
    else:
        filled     = int(df_steps[df_steps["IsFilled"]]["ID"].nunique())
        nok        = int(df_steps["IsNOK"].sum())
        vc         = df_steps["Value"].value_counts()
        ok         = int(vc.get("OK",  0))
        na         = int(vc.get("NA",  0))
        total_steps= len(df_steps)
        ctrl_dist  = df_steps["CtrlType"].value_counts().to_dict()
        cat_dist   = df_steps["Category"].value_counts().to_dict()
        year_dist  = df_steps.groupby("Year")["ID"].nunique().to_dict()

    rows = [
        ("── Extraction Info ──",    ""),
        ("WI Number",                wi_number),
        ("Extracted On",             datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("From Year",                str(from_year)),
        ("Record Limit",             str(limit) if limit > 0 else "All records"),
        ("",                         ""),
        ("── Record Counts ──",      ""),
        ("Total Records Pulled",     record_count),
        ("Filled Records",           filled),
        ("Empty / Draft Records",    record_count - filled),
        ("Parse Errors",             error_count),
        ("",                         ""),
        ("── Step Results ──",       ""),
        ("Total Step Rows (filled)", total_steps),
        ("OK Responses",             ok),
        ("NOK Responses",            nok),
        ("NA Responses",             na),
        ("",                         ""),
        ("── Records by Year ──",    ""),
    ] + [(f"  {y}", v) for y, v in sorted(year_dist.items())] + [
        ("",                         ""),
        ("── Control Type Mix ──",   ""),
    ] + [(f"  {k}", v) for k, v in sorted(ctrl_dist.items())] + [
        ("",                         ""),
        ("── Categories ──",         ""),
    ] + [(f"  {k}", v) for k, v in sorted(cat_dist.items())] + [
        ("",                         ""),
        ("── Output Files ──",       ""),
    ] + [(f"  File {i+1}", f) for i, f in enumerate(output_files)]

    return pd.DataFrame(rows, columns=["Metric", "Value"])


# ── EXCEL FORMATTING ──────────────────────────────────────────────────────────
def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def format_data_sheet(ws, col_widths: dict, result_col: str = None,
                      freeze_col: int = 5):
    """Format a data worksheet efficiently."""
    # Header row
    for cell in ws[1]:
        cell.font      = _font(bold=True, color="FFFFFF")
        cell.fill      = _fill(COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _border()
    ws.row_dimensions[1].height = 30

    # Data rows — batch process for speed
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = COLORS["alt_row"] if row_idx % 2 == 0 else COLORS["white"]
        for cell in row:
            cell.font      = _font(size=9)
            cell.border    = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if result_col and cell.column_letter == result_col:
                val = str(cell.value or "").strip()
                if   val == "OK":  cell.fill = _fill(COLORS["ok_bg"])
                elif val == "NOK":
                    cell.fill = _fill(COLORS["nok_bg"])
                    cell.font = _font(bold=True, color=COLORS["warn"], size=9)
                elif val == "NA":  cell.fill = _fill(COLORS["na_bg"])
                else:              cell.fill = _fill(bg)
            else:
                cell.fill = _fill(bg)

        ws.row_dimensions[row_idx].height = 15

    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes    = ws.cell(row=2, column=freeze_col + 1)
    ws.auto_filter.ref = ws.dimensions


def format_summary_sheet(ws):
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30
    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value or "")
            if val.startswith("──"):
                cell.font = _font(bold=True, color="FFFFFF", size=10)
                cell.fill = _fill(COLORS["section_bg"])
            else:
                cell.font = _font(size=10)
                cell.fill = _fill(COLORS["white"])
            cell.border    = _border()
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row[0].row].height = 18


STEPS_COL_WIDTHS = {
    "A": 8,  "B": 14, "C": 12, "D": 22, "E": 18,
    "F": 18, "G": 16, "H": 14, "I": 14, "J": 14,
    "K": 10, "L": 35, "M": 22, "N": 20, "O": 10,
    "P": 10, "Q": 30, "R": 8,  "S": 8,  "T": 8,
    "U": 8,  "V": 18, "W": 6,  "X": 8,
}

RAW_COL_WIDTHS = {
    "A": 8,  "B": 14, "C": 12, "D": 22, "E": 8,
    "F": 8,  "G": 10, "H": 10, "I": 22, "J": 20,
    "K": 35, "L": 18, "M": 18, "N": 10, "O": 22,
    "P": 30, "Q": 8,  "R": 8,  "S": 8,  "T": 8,
}


def apply_formatting(path: str):
    """Load workbook and apply all formatting. Handles large sheets efficiently."""
    log.info(f"Applying formatting to {os.path.basename(path)} ...")
    wb = load_workbook(path)

    if "Summary" in wb.sheetnames:
        format_summary_sheet(wb["Summary"])
        log.info("  Summary ✓")

    # Format all year sheets and the combined sheet
    for sheet_name in wb.sheetnames:
        if sheet_name in ("Summary", "Raw_Controls"):
            continue
        ws = wb[sheet_name]
        if ws.max_row <= 1:
            continue
        val_col = next(
            (c.column_letter for c in ws[1] if str(c.value) == "Value"), None
        )
        format_data_sheet(ws, STEPS_COL_WIDTHS,
                          result_col=val_col, freeze_col=5)
        log.info(f"  {sheet_name} ✓")

    if "Raw_Controls" in wb.sheetnames:
        ws = wb["Raw_Controls"]
        if ws.max_row > 1:
            format_data_sheet(wb["Raw_Controls"], RAW_COL_WIDTHS, freeze_col=4)
            log.info("  Raw_Controls ✓")

    wb.save(path)
    log.info(f"Formatting complete: {os.path.basename(path)}")


# ── SAFE FILE SAVE ────────────────────────────────────────────────────────────
def safe_save(writer_fn, path: str, retries: int = 3, delay: int = 5):
    """
    Save with retry logic. If the file is open in Excel,
    wait and retry rather than crashing.
    """
    for attempt in range(1, retries + 1):
        try:
            writer_fn(path)
            return path
        except PermissionError:
            if attempt < retries:
                log.warning(
                    f"  File is open in another application. "
                    f"Please close it. Retrying in {delay}s "
                    f"(attempt {attempt}/{retries})..."
                )
                time.sleep(delay)
            else:
                # Save to a new filename instead of failing
                base, ext = os.path.splitext(path)
                alt_path  = f"{base}_v{attempt}{ext}"
                log.warning(f"  Saving to alternate path: {alt_path}")
                writer_fn(alt_path)
                return alt_path


# ── OUTPUT ────────────────────────────────────────────────────────────────────
def safe_name(wi_number: str) -> str:
    return re.sub(r"[^\w]", "_", wi_number).strip("_")


def write_workbook(path: str, df_steps: pd.DataFrame, df_raw: pd.DataFrame,
                   df_summary: pd.DataFrame, years: list):
    """Write all sheets to a single workbook."""
    def _write(p):
        with pd.ExcelWriter(p, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM") as writer:
            df_summary.to_excel(writer, sheet_name="Summary",
                                index=False, header=False)

            # Year-by-year sheets
            for year in sorted(years):
                year_df = df_steps[df_steps["Year"] == year].copy()
                if year_df.empty:
                    continue
                year_df = year_df.drop(
                    columns=["Year", "Month", "IsFilled", "IsNOK"],
                    errors="ignore"
                )
                sheet_name = str(year)
                year_df.to_excel(writer, sheet_name=sheet_name, index=False)
                log.info(f"  Sheet '{sheet_name}': {len(year_df):,} rows")

            # Combined sheet (all years)
            combined = df_steps.drop(
                columns=["IsFilled", "IsNOK"], errors="ignore"
            )
            combined.to_excel(writer, sheet_name="All_Years", index=False)
            log.info(f"  Sheet 'All_Years': {len(combined):,} rows")

            # Raw controls
            if not df_raw.empty:
                df_raw_out = df_raw.drop(
                    columns=["Year", "GridRow", "GridCol"], errors="ignore"
                )
                df_raw_out.to_excel(writer, sheet_name="Raw_Controls", index=False)
                log.info(f"  Sheet 'Raw_Controls': {len(df_raw_out):,} rows")

    return safe_save(_write, path)


def save_outputs(df_steps: pd.DataFrame, df_raw: pd.DataFrame,
                 df_summary: pd.DataFrame, wi_number: str,
                 output_folder: str) -> list:
    """
    Save output. Splits into multiple files if row count exceeds MAX_ROWS_PER_FILE.
    Returns list of output file paths.
    """
    os.makedirs(output_folder, exist_ok=True)
    ts      = datetime.now().strftime("%Y%m%d_%H%M")
    base    = safe_name(wi_number)
    paths   = []

    if df_steps.empty:
        log.warning("No data to write.")
        return paths

    years = sorted(df_steps["Year"].dropna().unique().astype(int).tolist())
    total_rows = len(df_steps)

    if total_rows <= MAX_ROWS_PER_FILE:
        # Single file
        path = os.path.join(output_folder, f"{base}_{ts}.xlsx")
        path = write_workbook(path, df_steps, df_raw, df_summary, years)
        paths.append(path)
    else:
        # Split by year into separate files
        log.info(
            f"Row count {total_rows:,} exceeds limit {MAX_ROWS_PER_FILE:,}. "
            f"Splitting by year..."
        )
        for year in years:
            year_steps = df_steps[df_steps["Year"] == year].copy()
            year_raw   = df_raw[df_raw["Year"] == year].copy() if not df_raw.empty else pd.DataFrame()
            if year_steps.empty:
                continue
            path = os.path.join(output_folder, f"{base}_{year}_{ts}.xlsx")
            path = write_workbook(path, year_steps, year_raw, df_summary, [year])
            paths.append(path)
            log.info(f"  Saved {year}: {len(year_steps):,} rows → {os.path.basename(path)}")

    return paths


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Universal MOMS checklist extractor — any WI, any structure.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python extract.py --wi SIG/WR/PMW/0007
  python extract.py --wi SIG/WR/PMW/0007 --limit 1000
  python extract.py --wi SIG/WR/PMW/0007 --limit 0 --from_year 2021
  python extract.py --wi SIG/WR/PMW/0007 --output C:/Reports
        """
    )
    parser.add_argument("--wi",        required=True,
                        help="WI number e.g. SIG/WR/PMW/0007")
    parser.add_argument("--limit",     type=int, default=0,
                        help="Max records to fetch. 0 = all. Default: 0 (all)")
    parser.add_argument("--from_year", type=int, default=2021,
                        help="Only fetch records from this year onwards. Default: 2021")
    parser.add_argument("--output",    default="output",
                        help="Output folder. Default: output/")
    args = parser.parse_args()

    wi_number = args.wi.strip()

    log.info("=" * 60)
    log.info(f"  WI        : {wi_number}")
    log.info(f"  From Year : {args.from_year}")
    log.info(f"  Limit     : {args.limit or 'ALL'}")
    log.info(f"  Output    : {args.output}")
    log.info("=" * 60)

    # ── Fetch ─────────────────────────────────────────────────────────────────
    records = fetch_records(wi_number, args.limit, args.from_year)
    if not records:
        log.warning("No records found. Check WI number and date range.")
        sys.exit(0)

    # ── Parse ─────────────────────────────────────────────────────────────────
    all_steps, all_raw, errors = [], [], []

    for i, row in enumerate(records, 1):
        parsed = parse_record(row)
        if parsed["error"]:
            errors.append({"ID": row.ID, "Error": parsed["error"]})
            log.warning(f"  Skipped {row.ID}: {parsed['error']}")
        else:
            all_steps.extend(parsed["steps"])
            all_raw.extend(parsed["raw"])
        if i % 100 == 0:
            log.info(f"  Parsed {i:,} / {len(records):,} records ...")

    log.info(
        f"Parsing complete → "
        f"{len(all_steps):,} step rows | "
        f"{len(all_raw):,} raw rows | "
        f"{len(errors):,} errors"
    )

    # ── Build DataFrames ──────────────────────────────────────────────────────
    df_steps = pd.DataFrame(all_steps)
    df_raw   = pd.DataFrame(all_raw)
    df_steps, df_raw = postprocess(df_steps, df_raw)

    log.info(
        f"After dedup → "
        f"{len(df_steps):,} step rows | "
        f"{len(df_raw):,} raw rows"
    )

    # ── Save ──────────────────────────────────────────────────────────────────
    # Build summary with placeholder for output files (filled after save)
    df_summary = build_summary(
        df_steps, wi_number, args.limit, args.from_year,
        len(records), len(errors), []
    )

    output_paths = save_outputs(df_steps, df_raw, df_summary, wi_number, args.output)

    # Rebuild summary with actual filenames and apply formatting
    df_summary = build_summary(
        df_steps, wi_number, args.limit, args.from_year,
        len(records), len(errors),
        [os.path.basename(p) for p in output_paths]
    )

    for path in output_paths:
        # Update summary sheet in each file
        wb = load_workbook(path)
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            ws.delete_rows(1, ws.max_row)
            for r_idx, r in enumerate(df_summary.itertuples(index=False), 1):
                ws.cell(r_idx, 1, r[0])
                ws.cell(r_idx, 2, r[1])
            wb.save(path)

        apply_formatting(path)

    # ── Save errors ───────────────────────────────────────────────────────────
    if errors:
        err_path = os.path.join(
            args.output, f"{safe_name(wi_number)}_errors.csv"
        )
        pd.DataFrame(errors).to_csv(err_path, index=False)
        log.warning(f"Parse errors saved: {err_path}")

    # ── Final report ──────────────────────────────────────────────────────────
    nok_count = int(df_steps["IsNOK"].sum()) if not df_steps.empty else 0

    log.info("=" * 60)
    log.info(f"  WI Number      : {wi_number}")
    log.info(f"  Records        : {len(records):,}")
    log.info(f"  Step rows      : {len(df_steps):,}")
    log.info(f"  NOK responses  : {nok_count:,}")
    log.info(f"  Parse errors   : {len(errors):,}")
    for p in output_paths:
        size_mb = os.path.getsize(p) / 1024 / 1024
        log.info(f"  Output         : {p}  ({size_mb:.1f} MB)")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
    